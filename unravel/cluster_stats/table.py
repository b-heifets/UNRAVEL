#!/usr/bin/env python3

"""
Use ``cluster_table`` from UNRAVEL to summarize volumes of the top x regions and collapsing them into parent regions until a criterion is met.

Usage:
------
    cluster_table

Prereqs:
    ``cluster_index`` has been run. Run this command from the valid_clusters dir. <asterisk>cluster_info.txt in working dir.

Sorting by hierarchy and volume:
--------------------------------
Group by Depth: Starting from the earliest depth column, for each depth level:
   - Sum the volumes of all rows sharing the same region (or combination of regions up to that depth).
   - Sort these groups by their aggregate volume in descending order, ensuring larger groups are prioritized.

Sort Within Groups: Within each group created in step 1:
   - Sort the rows by their individual volume in descending order.

Maintain Grouping Order:
   - As we move to deeper depth levels, maintain the grouping and ordering established in previous steps (only adjusting the order within groups based on the new depth's aggregate volumes).

Note: 
    - CCFv3-2020_info.csv is in UNRAVEL/unravel/core/csvs/
    - It has columns: structure_id_path,very_general_region,collapsed_region_name,abbreviation,collapsed_region,other_abbreviation,other_abbreviation_defined,layer,sunburst
    - Alternatively, use CCFv3_info.csv (for gubra) or provide a custom CSV with the same columns.   
"""


import argparse
import openpyxl
import math
import numpy as np
import pandas as pd
from glob import glob
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font, Alignment
from rich import print
from rich.traceback import install

from unravel.core.argparse_utils import SuppressMetavar, SM
from unravel.core.config import Configuration
from unravel.core.utils import log_command, verbose_start_msg, verbose_end_msg


def parse_args():
    parser = argparse.ArgumentParser(formatter_class=SuppressMetavar)
    parser.add_argument('-vcd', '--val_clusters_dir', help='Path to the valid_clusters dir output from unravel.cluster_stats.index (else cwd)', action=SM)
    parser.add_argument('-t', '--top_regions', help='Number of top regions to output. Default: 4', default=4, type=int, action=SM)
    parser.add_argument('-pv', '--percent_vol', help='Percentage of the total volume the top regions must comprise [after collapsing]. Default: 0.8', default=0.8, type=float, action=SM)
    parser.add_argument('-csv', '--info_csv_path', help='CSV name or path/name.csv. Default: CCFv3-2020_info.csv', default='CCFv3-2020_info.csv', action=SM)
    parser.add_argument('-rgb', '--sunburst_rgbs', help='CSV name or path/name.csv. Default: sunburst_RGBs.csv', default='sunburst_RGBs.csv', action=SM)
    parser.add_argument('-v', '--verbose', help='Increase verbosity. Default: False', action='store_true', default=False)
    parser.epilog = __doc__
    return parser.parse_args()

# TODO: Correct font color for the volumes column. 'fiber tracts' is filled with white rather than the color of the fiber tracts
# TODO: 'CUL4, 5' is not filled with the color of the region. 


def fill_na_with_last_known(df):
    depth_columns = [col for col in df.columns if 'Depth' in col]

    # Fill NaN with the last known non-NaN value within each row for depth columns
    df_filled = df.copy()
    df_filled[depth_columns] = df_filled[depth_columns].fillna(method='ffill', axis=1)

    return df_filled

def sort_sunburst_hierarchy(df):
    """Sort the DataFrame by hierarchy and volume."""

    depth_columns = [col for col in df.columns if 'Depth' in col]
    volume_column = 'Volume_(mm^3)'

    # For each depth, process groups and sort
    for i, depth in enumerate(depth_columns):
        # Temporary DataFrame to hold sorting results for each depth
        sorted_partial = pd.DataFrame()
        
        # Identify unique groups (rows) up to the current depth (e.g, if Depth_2, then root, grey, CH)
        unique_groups = df[depth_columns[:i + 1]].drop_duplicates()
        
        for _, group_values in unique_groups.iterrows():
            # Filter rows belonging to the current group
            mask = (df[depth_columns[:i + 1]] == group_values).all(axis=1) # Boolean series to check which rows == group_values
            group_df = df[mask].copy() # Apply mask to df to get a df for each group (copy to avoid SettingWithCopyWarning)
            
            # Calculate aggregate volume for the group and add it as a new column
            group_df.loc[:, 'aggregate_volume'] = group_df[volume_column].sum()

            # Sort the group by individual volume
            group_df = group_df.sort_values(by=[volume_column], ascending=False)
            
            # Append sorted group to the partial result
            sorted_partial = pd.concat([sorted_partial, group_df], axis=0)

        # Replace df with the sorted_partial for the next iteration
        df = sorted_partial.drop(columns=['aggregate_volume'])

    return df

def undo_fill_with_original(df_sorted, df_original):
    # Ensure the original DataFrame has not been altered; otherwise, use a saved copy before any modifications
    depth_columns = [col for col in df_original.columns if 'Depth' in col]
    
    # Use the index to replace filled values with original ones where NaN existed
    for column in depth_columns:
        df_sorted[column] = df_original.loc[df_sorted.index, column] # Use the index to replace filled values with original ones where NaN existed
    
    return df_sorted

def can_collapse(df, depth_col):
    """
    Determine if regions in the specified depth column can be collapsed.
    Returns a DataFrame with regions that can be collapsed based on volume and count criteria.
    """
    volume_column = 'Volume_(mm^3)'

    # Group by the parent region and aggregate both volume and count (.e.g, if a parent has 3 children, the count is 3)
    subregion_aggregates = df.groupby(depth_col).agg({volume_column: ['sum', 'count']})
    subregion_aggregates.columns = ['Volume_Sum', 'Count']

    # Adjust the condition to check for both a volume threshold and a minimum count of subregions
    pooling_potential = subregion_aggregates[(subregion_aggregates['Volume_Sum'] > 0) & (subregion_aggregates['Count'] > 1)]

    return pooling_potential # DataFrame with regions that can be collapsed based on volume and count criteria (Depth_*, Volume_Sum, Count)

def collapse_hierarchy(df, verbose=False):
    volume_column = 'Volume_(mm^3)'
    depth_columns = [col for col in df.columns if 'Depth' in col]

    for depth_level in reversed(range(len(depth_columns))):
        depth_col = depth_columns[depth_level]
        if depth_level == 0: break  # Stop if we're at the top level

        # Identify regions that can be collapsed into their parent
        collapsible_regions = can_collapse(df, depth_col)

        # If collapsible_regions is not empty, proceed with collapsing
        if not collapsible_regions.empty:
            
            # For each collapsible region name, get the name and the aggregate volume
            for region, row in collapsible_regions.iterrows():

                aggregated_volume = row['Volume_Sum']

                # Collapse rows containing the region name and set the volume to the aggregate volume
                df.loc[df[depth_col] == region, volume_column] = aggregated_volume

                # Set the child region name to NaN in the depth column using the depth level
                child_depth_col = depth_columns[depth_level + 1] 
                df.loc[df[depth_col] == region, child_depth_col] = np.nan    

                # Also set any subsequent depth columns to NaN
                for subsequent_depth_col in depth_columns[depth_level + 2:]:
                    df.loc[df[depth_col] == region, subsequent_depth_col] = np.nan
               

                # Remove duplicate rows for the collapsed region
                df = df.drop_duplicates()

            return df


def calculate_top_regions(df, top_n, percent_vol_threshold, verbose=False):
    """
    Identify the top regions based on the dynamically collapsed hierarchy,
    ensuring they meet the specified percentage volume criterion.
    
    :param df_collapsed: DataFrame with the hierarchy collapsed where meaningful.
    :param top_n: The number of top regions to identify.
    :param percent_vol_threshold: The minimum percentage of total volume these regions should represent.
    :return: DataFrame with the top regions and their volumes if the criterion is met; otherwise, None.
    """
    # Get the total volume
    total_volume = df['Volume_(mm^3)'].sum()

    # Get top regions
    df_sorted = df.sort_values(by='Volume_(mm^3)', ascending=False).reset_index(drop=True)
    top_regions_df = df_sorted.head(top_n)

    # Sum the volumes of the top regions
    top_regions_volume = top_regions_df['Volume_(mm^3)'].sum()

    # Calculate the percentage of the total volume the top regions represent
    percent_vol = top_regions_volume / total_volume

    # Check if the top regions meet the percentage volume criterion
    if percent_vol >= percent_vol_threshold:
        return top_regions_df
    else:
        return None
    
def get_top_regions_and_percent_vols(sunburst_csv_path, top_regions, percent_vol, verbose=False):
    df = pd.read_csv(sunburst_csv_path)

    # Check if the DataFrame is empty print a message and return
    if df.empty:
        print(f'\n{sunburst_csv_path} is empty. Exiting...')
        import sys ; sys.exit()

    # Fill NaN values in the original DataFrame
    df_filled_na = fill_na_with_last_known(df.copy())

    # Sort the DataFrame by hierarchy and volume
    df_filled_sorted = sort_sunburst_hierarchy(df_filled_na)

    # Undo the fill with the original values
    df_final = undo_fill_with_original(df_filled_sorted, df)

    # Save the sorted DataFrame to a new CSV file
    sorted_parent_path = sunburst_csv_path.parent / '_sorted_sunburst_CSVs'
    sorted_parent_path.mkdir(parents=True, exist_ok=True)
    sorted_csv_name = str(sunburst_csv_path.name).replace('sunburst.csv', 'sunburst_sorted.csv')
    df_final.to_csv(sorted_parent_path / sorted_csv_name, index=False)

    # Attempt to calculate top regions, collapsing as necessary
    criteria_met = False
    while not criteria_met:
        top_regions_df = calculate_top_regions(df_final, top_regions, percent_vol, verbose)

        if top_regions_df is not None and not top_regions_df.empty: # If top regions are found
            criteria_met = True

            # If a top region contributes to less than 1% of the total volume, remove it
            total_volume = df_final['Volume_(mm^3)'].sum()
            top_regions_df = top_regions_df[top_regions_df['Volume_(mm^3)'] / total_volume > 0.01]

            # Initialize lists to hold the top region names and their aggregate volumes
            top_region_names_and_percent_vols = []

            # For each top region, get the region name and the aggregate volume
            for i, row in top_regions_df.iterrows():

                # Get the highest depth region name that is not NaN
                row_wo_volume = row[:10]
                region_name = row_wo_volume[::-1].dropna().iloc[0]

                # Calculate the percentage of the total volume the top regions represent
                aggregate_volume = row['Volume_(mm^3)']
                percent_vol = round(aggregate_volume / total_volume * 100)

                # Append the region name and the percentage volume to the list
                top_region_names_and_percent_vols.append(f'{region_name} ({percent_vol}%)')

            # Save the top regions DataFrame to a new CSV file
            top_regions_parent_path = sunburst_csv_path.parent / '_top_regions_for_each_cluster'
            top_regions_parent_path.mkdir(parents=True, exist_ok=True)
            top_regions_csv_name = str(sunburst_csv_path.name).replace('sunburst.csv', 'sunburst_top_regions.csv')
            top_regions_df.to_csv(top_regions_parent_path / top_regions_csv_name, index=False)
        else:
            # Attempt to collapse the hierarchy further
            df_final = collapse_hierarchy(df_final, verbose)
            if df_final.empty:
                break  # Exit if no further collapsing is possible

    return top_region_names_and_percent_vols, total_volume

# Function to create fill color based on the volume
def get_fill_color(value, max_value):
    # Apply a log transform to the volume to enhance visibility of smaller values
    # Adding 1 to the value to avoid log(0), and another 1 to max_value to ensure the max_value maps to 1 after log transform
    log_transformed_value = math.log10(value + 1)  
    log_transformed_max = math.log10(max_value + 1)
    normalized_value = log_transformed_value / log_transformed_max
    
    # Convert to a scale of 0-255 (for RGB values)
    rgb_value = int(normalized_value * 255)
    # Create fill color as a hex string
    fill_color = f"{rgb_value:02x}{rgb_value:02x}{rgb_value:02x}"
    
    # Set the font color to black if the fill color is light (more than 127 in RGB scale), otherwise white
    font_color = "000000" if rgb_value > 127 else "FFFFFF"
    
    return fill_color, font_color

@log_command
def main():
    install()
    args = parse_args()
    Configuration.verbose = args.verbose
    verbose_start_msg()

    # Find cluster_* dirs in the current dir
    valid_clusters_dir = Path(args.val_clusters_dir) if args.val_clusters_dir else Path.cwd()
    cluster_sunburst_csvs = valid_clusters_dir.glob('cluster_*_sunburst.csv')

    # Remove directories from the list
    cluster_sunburst_csvs = [f for f in cluster_sunburst_csvs if f.is_file()]

    # Generate dynamic column names based on args.top_regions
    column_names = ['Cluster'] + ['Volume'] + ['CoG'] + ['~Region'] + ['ID_Path'] + [f'Top_Region_{i+1}' for i in range(args.top_regions)]

    # Load the *cluster_info.txt file from the parent dir to get the cluster Centroid of Gravity (CoG)
    cluster_info_txt_parent = Path(args.val_clusters_dir).parent if args.val_clusters_dir else Path.cwd()
    cluster_info_files = list(cluster_info_txt_parent.glob('*cluster_info.txt'))
    
    # If no cluster_info.txt file is found, exit
    if not cluster_info_files:
        print(f'\n    [red]No *cluster_info.txt file found in {cluster_info_txt_parent}. Exiting...')
        return
    else:
        cluster_info_file = cluster_info_files[0] # first match
 
    # Load the cluster info file
    cluster_info_df = pd.read_csv(cluster_info_file, sep='\t', header=None)  # Assuming tab-separated values; adjust if different

    # Reverse the row order of only the first column (excluding the header)
    first_column_name = cluster_info_df.columns[0]
    reversed_data = cluster_info_df[first_column_name].iloc[1:].iloc[::-1].reset_index(drop=True) # Reverse the data 
    new_first_column = pd.concat([cluster_info_df[first_column_name].iloc[:1], reversed_data]) # Concat header w/ reversed data
    cluster_info_df[first_column_name] = new_first_column.values # Assign the new column back to the DataFrame

    # Get the CoG for each cluster (values in last three columns of each row)
    CoGs = cluster_info_df.iloc[:, -3:].values

    # Convert to this format: 'x,y,z'
    CoGs = [','.join(map(str, CoG)) for CoG in CoGs]

    # Create a dict w/ the first column as keys and the CoG as values
    cluster_CoGs = dict(zip(cluster_info_df[first_column_name], CoGs))

    # Create an empty DataFrame with the column names
    top_regions_and_percent_vols_df = pd.DataFrame(columns=column_names)

    # Specify the column names you want to load
    columns_to_load = ['abbreviation', 'general_region',  'structure_id_path']

    # Load the specified columns from the CSV with CCFv3 info
    if args.info_csv_path == 'CCFv3_info.csv' or args.info_csv_path == 'CCFv3-2020_info.csv': 
        ccfv3_info_df = pd.read_csv(Path(__file__).parent.parent / 'core' / 'csvs' / args.info_csv_path, usecols=columns_to_load)
    else:
        ccfv3_info_df = pd.read_csv(args.info_csv_path, usecols=columns_to_load)

    # For each cluster directory
    for cluster_sunburst_csv in cluster_sunburst_csvs:

        # Get the cluster number from the file name 'cluster_*_sunburst.csv'
        cluster_num = str(cluster_sunburst_csv.name).split('_')[1]

        # Get the CoG string for the current cluster from the dictionary
        cog_string = cluster_CoGs.get(cluster_num) if cluster_CoGs.get(cluster_num) else "Not found"

        # Get the top regions and their percentage volumes for the current cluster
        top_regions_and_percent_vols, cluster_volume = get_top_regions_and_percent_vols(cluster_sunburst_csv, args.top_regions, args.percent_vol, args.verbose)

        # Get the top region
        top_region = top_regions_and_percent_vols[0].split(' ')[0]

        # Lookup the general_regio and structure_id_path in the DataFrame using the abbreviation for the top region
        general_region = ccfv3_info_df.loc[ccfv3_info_df['abbreviation'] == top_region, 'general_region'].values
        id_path = ccfv3_info_df.loc[ccfv3_info_df['abbreviation'] == top_region, 'structure_id_path'].values

        # Since there could be multiple matches, we take the first one or a default value if not found
        general_region = general_region[0] if len(general_region) > 0 else "General region not found"
        id_path = id_path[0] if len(id_path) > 0 else "ID path not found"

        # Ensure the list has the exact number of top regions (pad with None if necessary)
        padded_top_regions = (list(top_regions_and_percent_vols) + [None] * args.top_regions)[:args.top_regions]

        # Prepare the row data, including placeholders for 'Volume', 'CoG', '~Region', and top regions
        row_data = [cluster_num, cluster_volume, cog_string, general_region, id_path] + padded_top_regions   

        # Ensure column_names matches the structure of row_data
        column_names = ['Cluster', 'Volume', 'CoG', '~Region', 'ID_Path'] + [f'Top_Region_{i+1}' for i in range(args.top_regions)]

        # Create a temporary DataFrame for the current cluster's data
        temp_df = pd.DataFrame([row_data], columns=column_names)

        # Concatenate the temporary DataFrame with the main DataFrame
        top_regions_and_percent_vols_df = pd.concat([top_regions_and_percent_vols_df, temp_df], ignore_index=True)

    # Sort the DataFrame by the 'ID_Path' column in descending order
    top_regions_and_percent_vols_df = top_regions_and_percent_vols_df.sort_values(by='ID_Path', ascending=False)

    # Drop the 'ID_Path' column
    top_regions_and_percent_vols_df = top_regions_and_percent_vols_df.drop(columns=['ID_Path'])

    # Convert the 'Volume' column to 4 decimal places
    top_regions_and_percent_vols_df['Volume'] = top_regions_and_percent_vols_df['Volume'].round(4)
    print(f'\nThe top regions and their percentage volumes for each cluster:')
    print(f'\n{top_regions_and_percent_vols_df.to_string(index=False)}\n')

    # Load csv with RGB values 
    if args.sunburst_rgbs == 'sunburst_RGBs.csv': 
        sunburst_RGBs_df = pd.read_csv(Path(__file__).parent.parent / 'core' / 'csvs' / 'sunburst_RGBs.csv', header=None)
    else:
        sunburst_RGBs_df = pd.read_csv(args.sunburst_rgbs, header=None)

    # Parse the dataframe to get a dictionary of region names and their corresponding RGB values
    rgb_values = {}
    for index, row in sunburst_RGBs_df.iterrows():
        # Format is 'region_name: rgb(r,g,b)'
        region, rgb = row[0].split(': rgb(')
        r, g, b = map(int, rgb.strip(')').split(','))
        rgb_values[region] = (r, g, b)

    # Create an Excel workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write each row of the DataFrame to the worksheet and color cells for the top regions
    for region in dataframe_to_rows(top_regions_and_percent_vols_df, index=False, header=True):
        ws.append(region)

        for i in range(args.top_regions):

            # Find the region name without the percentage to match the RGB values
            top_region_column_num = 4 + i

            region_key = None
            if region[top_region_column_num] is not None:
                region_key = region[top_region_column_num].split(' ')[0]

            # Apply the color to the cell if it matches one of the RGB values
            if region_key in rgb_values:
                # Convert the RGB to a hex string
                rgb = rgb_values[region_key]
                hex_color = "{:02x}{:02x}{:02x}".format(rgb[0], rgb[1], rgb[2])
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                top_region_column_num = 5 + i
                ws.cell(row=ws.max_row, column=top_region_column_num).fill = fill
                ws.cell(row=ws.max_row, column=top_region_column_num).border = thin_border
            elif region_key is None:
                hex_color = "{:02x}{:02x}{:02x}".format(100, 100, 100) # Grey
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                top_region_column_num = 5 + i
                ws.cell(row=ws.max_row, column=top_region_column_num).fill = fill
                ws.cell(row=ws.max_row, column=top_region_column_num).border = thin_border

    # Insert a new row at the top
    ws.insert_rows(1)

    # Insert a new column at the left
    ws.insert_cols(1)            

    # Adjust the column width to fit the content
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if max_length > 0:
            adjusted_width = max_length + 2  # Add 2 for a little extra padding
            column_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Fill cells in first column with white
    for cell in ws['A']:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in ninth column with white
    column = 6 + args.top_regions
    for cell in ws[get_column_letter(column)]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in first row with white
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in last row with white
    num_rows = top_regions_and_percent_vols_df.shape[0] + 3
    for cell in ws[num_rows]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Center align the content
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format column C such that the brightness of the fill is proportional to the volume / total volume and the text brightness is inversely proportional to the fill brightness
    # total_volume = top_regions_and_percent_vols_df['Volume'].sum()
    volumes = top_regions_and_percent_vols_df['Volume'].tolist()
    max_volume = max(volumes)  # The largest volume

    for row, volume in enumerate(volumes, start=3):  # Starting from row 3 (C3)
        # Calculate the color based on the volume
        fill_color, font_color = get_fill_color(volume, max_volume)
        
        # Get the cell at column C and the current row
        cell = ws.cell(row=row, column=3)  # Column 3 corresponds to column C
        
        # Set the fill color based on the calculated brightness
        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Set the font color based on the inverse of the fill color's brightness
        cell.font = Font(color=font_color)

    # Apply a thin border style to cells with content
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:  # If the cell has content
                cell.border = thin_border
                cell.font = Font(name='Arial', size=11)

    # Apply the font to the header row
    header_font = Font(name='Arial', bold=True)
    for cell in ws['2:2']:
        cell.font = header_font

    # Make column B bold
    for cell in ws['B']:
        cell.font = Font(name='Arial', bold=True)
    
    # Additional step to ensure cells from column F onwards are black
    for row in ws.iter_rows(min_col=6): 
        for cell in row:
            if cell.font:  # If the cell already has font settings applied
                cell.font = Font(name='Arial', size=cell.font.size, bold=cell.font.bold, color='FF000000')
            else:
                cell.font = Font(name='Arial', color='FF000000')

    # Iterate through the cells and merge cells with the same value in column 5
    current_region = None
    first_row = None

    # Adjusted min_row to 2 and min_col/max_col to merge_column because of the added padding row and column
    merge_column = 5
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=merge_column, max_col=merge_column):
        cell = row[0]  # row[0] since we're only looking at one column, and iter_rows yields tuples
        if cell.value != current_region:
            # If the cell value changes, merge the previous cells if there are more than one with the same value
            if first_row and first_row < cell.row - 1:
                ws.merge_cells(start_row=first_row, start_column=merge_column, end_row=cell.row - 1, end_column=merge_column)
            # Update the current region and reset the first_row to the current cell's row
            current_region = cell.value
            first_row = cell.row

    # After the loop, check and merge the last set of cells if needed
    if first_row and first_row < ws.max_row:
        ws.merge_cells(start_row=first_row, start_column=merge_column, end_row=ws.max_row - 1, end_column=merge_column)

    for cell in ws['E']:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Save the workbook to a file
    excel_file_path = valid_clusters_dir / f'{valid_clusters_dir.parent.name}_valid_clusters_table.xlsx'
    wb.save(excel_file_path)
    print(f"Excel file saved at {excel_file_path}")
    print(f"\nBrighter cell fills in the 'Volume' column represent larger volumes (log(10) scaled and normalized to the max volume).")

    # Get the anatomically sorted list of cluster IDs and save it to a .txt file
    valid_cluster_ids = top_regions_and_percent_vols_df['Cluster'].tolist()
    valid_cluster_ids_str = ' '.join(map(str, valid_cluster_ids)) + '\n'
    with open(valid_clusters_dir / 'valid_cluster_IDs_sorted_by_anatomy.txt', 'w') as f:
        f.write(valid_cluster_ids_str)

    verbose_end_msg()
    

if __name__ == '__main__':
    main()