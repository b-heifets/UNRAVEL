#!/usr/bin/env python3

"""
Use ``cluster_legend`` from UNRAVEL to summarize regional abbreviations from <asterisk>_valid_clusters_table.xlsx files.

Usage:
------
    cluster_legend

Inputs:
    <asterisk>_valid_clusters_table.xlsx files in the working directory output from ``cluster_table``

Outputs:
    legend.xlsx

Note: 
    - CCFv3-2020_info.csv is in UNRAVEL/unravel/core/csvs/
    - It has columns: structure_id_path,very_general_region,collapsed_region_name,abbreviation,collapsed_region,other_abbreviation,other_abbreviation_defined,layer,sunburst
    - Alternatively, use CCFv3_info.csv (for gubra) or provide a custom CSV with the same columns.
"""

import argparse
from glob import glob
from pathlib import Path
import openpyxl
import pandas as pd
from rich import print
from rich.traceback import install
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import Alignment

from unravel.core.argparse_utils import SuppressMetavar, SM
from unravel.core.config import Configuration
from unravel.core.utils import log_command, verbose_start_msg, verbose_end_msg


def parse_args():
    parser = argparse.ArgumentParser(formatter_class=SuppressMetavar)
    parser.add_argument('-p', '--path', help='Path to the directory containing the *_valid_clusters_table.xlsx files. Default: current working directory', action=SM)
    parser.add_argument('-csv', '--csv_path', help='CSV name or path/name.csv. Default: CCFv3-2020_info.csv', default='CCFv3-2020_info.csv', action=SM)
    parser.add_argument('-v', '--verbose', help='Increase verbosity. Default: False', action='store_true', default=False)
    parser.epilog = __doc__
    return parser.parse_args()


def extract_unique_regions_from_file(file_path):
    # Load the Excel file into a DataFrame, specifying that the header is in the second row (index 1)
    df = pd.read_excel(file_path, header=1, engine='openpyxl')
    
    unique_regions = set()

    # Iterate through columns and find those that start with "Top_Region"
    for col in df.columns:
        if col.startswith("Top_Region"):
            # Extract the regions from the column, remove percent volumes and add to the set
            unique_regions.update(df[col].dropna().apply(lambda x: ' '.join(x.split()[:-1]))) 

    return unique_regions

def apply_rgb_to_cell(ws, df_w_rgb, col_w_labels, col_num):
    """Apply RGB values to cells in the worksheet. 
    
    Parameters: ws (openpyxl worksheet), df_w_rgb (DataFrame with RGB values), col_w_labels (column with region labels), region (region name), col_num (column number to apply the RGB values starting from 0)"""
    for row in ws.iter_rows(min_row=3, min_col=2, max_col=4, max_row=ws.max_row-1):
        region_cell = row[col_num]

        r = df_w_rgb.loc[df_w_rgb[col_w_labels] == region_cell.value, 'R'].values[0]
        g = df_w_rgb.loc[df_w_rgb[col_w_labels] == region_cell.value, 'G'].values[0]
        b = df_w_rgb.loc[df_w_rgb[col_w_labels] == region_cell.value, 'B'].values[0]
        hex_color = "{:02x}{:02x}{:02x}".format(r, g, b)
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
        region_cell.fill = fill

@log_command
def main():
    install()
    args = parse_args()
    Configuration.verbose = args.verbose
    verbose_start_msg()

    path = Path(args.path) if args.path else Path.cwd()

    # Find cluster_* dirs in the current dir
    xlsx_files = path.glob('*_valid_clusters_table.xlsx')

    # Filter out files starting with '~$'
    xlsx_files = [f for f in xlsx_files if not str(f).split('/')[-1].startswith('~$')]

    # Filter out files that start with legend
    xlsx_files = [f for f in xlsx_files if not str(f).split('/')[-1].startswith('legend')]

    if xlsx_files == []:
        print("\n    [red1]No *_valid_clusters_table.xlsx files found in the specified directory. Exiting...\n")
        import sys ; sys.exit()
    else:
        print(f'\nProcessing:')
        for file in xlsx_files:
            print(f'    {file}')

    # Initialize a set to store unique regions from all files, accounting for headers in the second row
    all_unique_regions = set() # Using a set to avoid duplicates

    # Iterate through each file and extract unique regions from each file
    for file_path in xlsx_files:
        unique_regions = extract_unique_regions_from_file(file_path)
        all_unique_regions.update(unique_regions)

    # Convert the set to a sorted list for easier reading
    all_unique_regions = sorted(list(all_unique_regions))

    if len(all_unique_regions) == 0:
        print("\n    [red1]No regions found in the xlsx files. Headers expected in the second row. Exiting...\n")
        import sys ; sys.exit()

    print(f'\nRegions: {all_unique_regions}\n')

    # Specify the column names you want to load
    columns_to_load = ['structure_id_path', 'very_general_region',  'collapsed_region_name', 'abbreviation', 'collapsed_region', 'other_abbreviation', 'other_abbreviation_defined', 'layer', 'sunburst']

    # Load the specified columns from the CSV with CCFv3 info
    if args.csv_path == 'CCFv3_info.csv' or args.csv_path == 'CCFv3-2020_info.csv': 
        ccfv3_info_df = pd.read_csv(Path(__file__).parent.parent / 'core' / 'csvs' / args.csv_path, usecols=columns_to_load)
    else:
        ccfv3_info_df = pd.read_csv(args.csv_path, usecols=columns_to_load)

    # Creat a dictionary to hold the mappings for the region abbreviation to collapsed region abbreviation
    abbreviation_to_collapsed_dict = dict(zip(ccfv3_info_df['abbreviation'], ccfv3_info_df['collapsed_region']))

    # Now collapse the regions in the unique_regions set
    unique_regions_collapsed = {abbreviation_to_collapsed_dict.get(region, region) for region in all_unique_regions} 
    unique_regions_collapsed = sorted(list(unique_regions_collapsed))
    print(f'{unique_regions_collapsed=}\n')

    # If a region in all_unique_regions has a digit in it, check if the 'layer' column is defined for it. Then, add unique layers to a set
    layers_set = set()
    for region in all_unique_regions:
        if any(char.isdigit() for char in region):
            layer = ccfv3_info_df.loc[ccfv3_info_df['abbreviation'] == region, 'layer'].values
            if len(layer) > 0:
                layers_set.add(str(layer[0]))  # Convert float to string

    # Sort the layers
    layers_set = sorted(list(layers_set))
    layers_set = [layer for layer in layers_set if str(layer) != 'nan']

    # Get all regions with digits that are not defined as layers
    other_regions_w_digits = [
        region for region in all_unique_regions
        if any(char.isdigit() for char in region) and not ccfv3_info_df[ccfv3_info_df['abbreviation'] == region]['layer'].notna().any()
    ]

    # Print the cortical layers and any regions with digits that are not defined as layers
    if len(other_regions_w_digits) > 0:
        print(f"Numbers ({layers_set}) = cortical layers (with these exceptions {other_regions_w_digits})\n")
    else:
        print(f"Numbers ({layers_set}) = cortical layers\n")

    # For regions in all_unique_regions, determine abbreviations to offload from the table (i.e., abbreviations mentioned in 'other_abbreviation' and defined in 'other_abbreviation_defined')
    list_of_regions_w_other_abbreviation_in_all_unique_regions = [region for region in all_unique_regions if ccfv3_info_df.loc[ccfv3_info_df['abbreviation'] == region, 'other_abbreviation'].notna().any()]

    # Initialize an empty dictionary to hold the mapping of other_abbreviations to their definitions
    other_abbreviation_to_definitions = {}

    for region in list_of_regions_w_other_abbreviation_in_all_unique_regions:
        # Extract 'other_abbreviation' and 'other_abbreviation_defined' for the current region
        rows = ccfv3_info_df[ccfv3_info_df['abbreviation'] == region]
        for _, row in rows.iterrows():
            other_abbreviation = row['other_abbreviation']
            other_abbreviation_defined = row['other_abbreviation_defined']
            
            if pd.notna(other_abbreviation) and pd.notna(other_abbreviation_defined):
                # Initialize the set for this abbreviation if it doesn't exist
                if other_abbreviation not in other_abbreviation_to_definitions:
                    other_abbreviation_to_definitions[other_abbreviation] = set()
                
                # Add the current definition to the set of definitions for this abbreviation
                other_abbreviation_to_definitions[other_abbreviation].add(other_abbreviation_defined)

    # Convert sets to strings with " or " as the separator
    for abbreviation, definitions_set in other_abbreviation_to_definitions.items():
        other_abbreviation_to_definitions[abbreviation] = " or ".join(definitions_set)

    # Sort the dictionary by key
    other_abbreviation_to_definitions = dict(sorted(other_abbreviation_to_definitions.items()))
    print(f'Additional abbreviations not shown in the region abbreviation legend (SI table): {other_abbreviation_to_definitions}')

    # Get the 'very_general_region' column from the CCFv3_info.csv file and use it to get the 'very_general_region' for each region in unique_regions_collapsed
    very_general_region_dict = dict(zip(ccfv3_info_df['collapsed_region'], ccfv3_info_df['very_general_region']))
    very_general_regions = [very_general_region_dict.get(region, '') for region in unique_regions_collapsed]

    # If the same string is in the 'very_general_regions' list and the 'unique_regions_collapsed' list, remove it from both at the same index
    for i, region in enumerate(unique_regions_collapsed):
        if very_general_regions[i] == region:
            very_general_regions[i] = ''
            unique_regions_collapsed[i] = ''

    # Reset the indices of the lists
    very_general_regions = [region for region in very_general_regions if region != '']
    unique_regions_collapsed = [region for region in unique_regions_collapsed if region != '']

    # Make a dataframe with the 'very_general_regions' and 'unique_regions_collapsed' lists
    legend_df = pd.DataFrame({'Region': very_general_regions, 'Abbrev.': unique_regions_collapsed})

    # Add the 'Subregion' column to the dataframe
    legend_df['Subregion'] = [ccfv3_info_df.loc[ccfv3_info_df['collapsed_region'] == region, 'collapsed_region_name'].values[0] for region in unique_regions_collapsed]

    # Add the 'structure_id_path' column to the dataframe
    legend_df['structure_id_path'] = [ccfv3_info_df.loc[ccfv3_info_df['collapsed_region'] == region, 'structure_id_path'].values[0] for region in unique_regions_collapsed]

    # Sort the dataframe by the 'structure_id_path' column in descending order
    legend_df.sort_values(by='structure_id_path', ascending=False, inplace=True)

    # Reset the index of the dataframe
    legend_df.reset_index(drop=True, inplace=True)

    # Drop the 'structure_id_path' column
    legend_df.drop(columns='structure_id_path', inplace=True)

    # Create an Excel workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write the dataframe to the worksheet
    for region in dataframe_to_rows(legend_df, index=False, header=True):
        ws.append(region)

    # Apply a thin border style to cells with content
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.border = thin_border
                cell.font = Font(name='Arial', size=11)

    # Apply the font to the header row
    header_font = Font(name='Arial', bold=True)
    for cell in ws['1:1']:
        cell.font = header_font

    # Make the first column bold
    for cell in ws['A']:
        cell.font = Font(name='Arial', bold=True)

    # Insert a new row at the top
    ws.insert_rows(1)

    # Insert a new column at the left
    ws.insert_cols(1)        

    # Fill cells in first column with white
    for cell in ws['A']:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in fifth column with white
    for cell in ws['E']:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in first row with white
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in last row with white
    num_rows = legend_df.shape[0] + 3
    for cell in ws[num_rows]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Adjust the column width to fit the content
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if max_length > 0:
            adjusted_width = max_length + 2
            column_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    # # Add columns for R, G, and B values to ccfv3_info_df
    ccfv3_info_df[['R', 'G', 'B']] = ccfv3_info_df['sunburst'].str.extract(r'rgb\((\d+),(\d+),(\d+)\)')
    ccfv3_info_df[['R', 'G', 'B']] = ccfv3_info_df[['R', 'G', 'B']].apply(pd.to_numeric)

    # Apply RGB values to cells
    apply_rgb_to_cell(ws, ccfv3_info_df, 'very_general_region', 0)
    apply_rgb_to_cell(ws, ccfv3_info_df, 'collapsed_region', 1)
    apply_rgb_to_cell(ws, ccfv3_info_df, 'collapsed_region_name', 2)

    # Iterate through the cells and merge cells with the same value in column B
    current_region = None
    first_row = None

    # Adjusted min_row to 2 and min_col/max_col to merge_column because of the added padding row and column
    merge_column = 2
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, min_col=merge_column, max_col=merge_column):
        cell = row[0]  # row[0] since we're only looking at one column, and iter_rows yields tuples
        if cell.value != current_region:
            # If the cell value changes, merge the previous cells if there are more than one with the same value
            if first_row and first_row < cell.row - 1:
                ws.merge_cells(start_row=first_row, start_column=merge_column, end_row=cell.row - 1, end_column=merge_column)
                # After merging, we need to set the alignment for the merged cells
                merged_cell = ws.cell(row=first_row, column=merge_column)
                merged_cell.alignment = Alignment(vertical='center')
            # Update the current region and reset the first_row to the current cell's row
            current_region = cell.value
            first_row = cell.row

    # After the loop, check and merge the last set of cells if needed
    if first_row and first_row < ws.max_row:
        ws.merge_cells(start_row=first_row, start_column=merge_column, end_row=ws.max_row - 1, end_column=merge_column)
        # Align the last merged cell as well
        merged_cell = ws.cell(row=first_row, column=merge_column)
        merged_cell.alignment = Alignment(vertical='center')

    # Center align the content
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ensure that fonts are black
    for row in ws.iter_rows(min_col=2): 
        for cell in row:
            if cell.font:  # If the cell already has font settings applied
                cell.font = Font(name='Arial', size=cell.font.size, bold=cell.font.bold, color='FF000000')
            else:
                cell.font = Font(name='Arial', color='FF000000')

    # Save the workbook to a file
    excel_file_path = path / 'legend.xlsx'
    wb.save(excel_file_path)

    # Save text for figure legend 
    fig_legend_txt = path / "fig_legend.txt"
    with open(fig_legend_txt, "w") as file:
        file.write(f'\n{all_unique_regions=}\n')
        file.write(f'\n{unique_regions_collapsed=}\n')
        if len(other_regions_w_digits) > 0:
            file.write(f"\nNumbers ({layers_set}) = cortical layers (with these exceptions {other_regions_w_digits})\n")
        else:
            file.write(f"\nNumbers ({layers_set}) = cortical layers\n")
        file.write(f'\nAdditional abbreviations not shown in the region abbreviation legend (SI table): {other_abbreviation_to_definitions}\n')

    verbose_end_msg()


if __name__ == '__main__':
    main()