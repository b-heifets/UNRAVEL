#!/usr/bin/env python3

import argparse
from glob import glob
from pathlib import Path
from matplotlib.colors import hex2color
import numpy as np
import openpyxl
import pandas as pd
from argparse_utils import SuppressMetavar, SM
from rich import print
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Font
from openpyxl.styles import Alignment

def parse_args():
    parser = argparse.ArgumentParser(description='''Summarize volumes of the top x regions and collapsing them into parent regions until a criterion is met.''',
                                     formatter_class=SuppressMetavar)
    parser.add_argument('-t', '--top_regions', help='Number of top regions to output. Default: 4', default=4, type=int, action=SM)
    parser.add_argument('-p', '--percent_vol', help='Percentage of the total volume the top regions must comprise [after collapsing]. Default: 0.8', default=0.8, type=float, action=SM)
    parser.add_argument('-v', '--verbose', help='Increase verbosity. Default: False', action='store_true', default=False)
    parser.epilog = """Example usage:    valid_clusters_table.py

Prerequisites: valid_cluster_index.sh has been run. Run this script from the <valid_clusters> dir. *cluster_info.txt is assumed to be in its parent dir.

Sorting by heirarchy and volume: 
1) Group by Depth: Starting from the earliest depth column, for each depth level:
       - Sum the volumes of all rows sharing the same region (or combination of regions up to that depth).
       - Sort these groups by their aggregate volume in descending order, ensuring larger groups are prioritized.

2) Sort Within Groups: Within each group created in step 1:
       - Sort the rows by their individual volume in descending order.

3) Maintain Grouping Order: 
       - As we move to deeper depth levels, maintain the grouping and ordering established in previous steps
       (only adjusting the order within groups based on the new depth's aggregate volumes).  
    
"""
    return parser.parse_args()


def fill_na_with_last_known(df):
    depth_columns = [col for col in df.columns if 'Depth' in col]

    # Fill NaN with the last known non-NaN value within each row for depth columns
    df_filled = df.copy()
    df_filled[depth_columns] = df_filled[depth_columns].fillna(method='ffill', axis=1)

    return df_filled

def sort_sunburst_hierarchy(df):
    """Sort the DataFrame by hierarchy and volume."""

    depth_columns = [col for col in df.columns if 'Depth' in col]
    volume_column = 'Volume_(mm^3)'

    # For each depth, process groups and sort
    for i, depth in enumerate(depth_columns):
        # Temporary DataFrame to hold sorting results for each depth
        sorted_partial = pd.DataFrame()
        
        # Identify unique groups (rows) up to the current depth (e.g, if Depth_2, then root, grey, CH)
        unique_groups = df[depth_columns[:i + 1]].drop_duplicates()
        
        for _, group_values in unique_groups.iterrows():
            # Filter rows belonging to the current group
            mask = (df[depth_columns[:i + 1]] == group_values).all(axis=1) # Boolean series to check which rows == group_values
            group_df = df[mask].copy() # Apply mask to df to get a df for each group (copy to avoid SettingWithCopyWarning)
            
            # Calculate aggregate volume for the group and add it as a new column
            group_df.loc[:, 'aggregate_volume'] = group_df[volume_column].sum()

            # Sort the group by individual volume
            group_df = group_df.sort_values(by=[volume_column], ascending=False)
            
            # Append sorted group to the partial result
            sorted_partial = pd.concat([sorted_partial, group_df], axis=0)

        # Replace df with the sorted_partial for the next iteration
        df = sorted_partial.drop(columns=['aggregate_volume'])

    return df

def undo_fill_with_original(df_sorted, df_original):
    # Ensure the original DataFrame has not been altered; otherwise, use a saved copy before any modifications
    depth_columns = [col for col in df_original.columns if 'Depth' in col]
    
    # Use the index to replace filled values with original ones where NaN existed
    for column in depth_columns:
        df_sorted[column] = df_original.loc[df_sorted.index, column] # Use the index to replace filled values with original ones where NaN existed
    
    return df_sorted

def can_collapse(df, depth_col):
    """
    Determine if regions in the specified depth column can be collapsed.
    Returns a DataFrame with regions that can be collapsed based on volume and count criteria.
    """
    volume_column = 'Volume_(mm^3)'

    # Group by the parent region and aggregate both volume and count (.e.g, if a parent has 3 children, the count is 3)
    subregion_aggregates = df.groupby(depth_col).agg({volume_column: ['sum', 'count']})
    subregion_aggregates.columns = ['Volume_Sum', 'Count']

    # Adjust the condition to check for both a volume threshold and a minimum count of subregions
    pooling_potential = subregion_aggregates[(subregion_aggregates['Volume_Sum'] > 0) & (subregion_aggregates['Count'] > 1)]

    return pooling_potential # DataFrame with regions that can be collapsed based on volume and count criteria (Depth_*, Volume_Sum, Count)

def collapse_hierarchy(df, verbose=False):
    volume_column = 'Volume_(mm^3)'
    depth_columns = [col for col in df.columns if 'Depth' in col]

    for depth_level in reversed(range(len(depth_columns))):
        depth_col = depth_columns[depth_level]
        if depth_level == 0: break  # Stop if we're at the top level

        # Identify regions that can be collapsed into their parent
        collapsible_regions = can_collapse(df, depth_col)

        if verbose:
            print(f'\n{collapsible_regions=}\n')

        # If collapsible_regions is not empty, proceed with collapsing
        if not collapsible_regions.empty:
            
            # For each collapsible region name, get the name and the aggregate volume
            for region, row in collapsible_regions.iterrows():

                aggregated_volume = row['Volume_Sum']

                # Collapse rows containing the region name and set the volume to the aggregate volume
                df.loc[df[depth_col] == region, volume_column] = aggregated_volume

                # Set the child region name to NaN in the depth column using the depth level
                child_depth_col = depth_columns[depth_level + 1] 
                df.loc[df[depth_col] == region, child_depth_col] = np.nan    

                # Also set any subsequent depth columns to NaN
                for subsequent_depth_col in depth_columns[depth_level + 2:]:
                    df.loc[df[depth_col] == region, subsequent_depth_col] = np.nan
               

                # Remove duplicate rows for the collapsed region
                df = df.drop_duplicates()

            return df


def calculate_top_regions(df, top_n, percent_vol_threshold, verbose=False):
    """
    Identify the top regions based on the dynamically collapsed hierarchy,
    ensuring they meet the specified percentage volume criterion.
    
    :param df_collapsed: DataFrame with the hierarchy collapsed where meaningful.
    :param top_n: The number of top regions to identify.
    :param percent_vol_threshold: The minimum percentage of total volume these regions should represent.
    :return: DataFrame with the top regions and their volumes if the criterion is met; otherwise, None.
    """
    # Get the total volume
    total_volume = df['Volume_(mm^3)'].sum()

    # Get top regions
    df_sorted = df.sort_values(by='Volume_(mm^3)', ascending=False).reset_index(drop=True)
    top_regions_df = df_sorted.head(top_n)

    # Sum the volumes of the top regions
    top_regions_volume = top_regions_df['Volume_(mm^3)'].sum()

    # Calculate the percentage of the total volume the top regions represent
    percent_vol = top_regions_volume / total_volume

    if verbose:
        print(f'The percentage of the total volume the top {top_n} regions represents: {percent_vol}\n')

    # Check if the top regions meet the percentage volume criterion
    if percent_vol >= percent_vol_threshold:
        return top_regions_df
    else:
        return None
    
def get_top_regions_and_percent_vols(cluster_dir, top_regions, percent_vol, verbose=False):
    # Load the CSV file into a DataFrame
    cluster_path = Path(cluster_dir).resolve()
    csv_files = glob(f'{cluster_path}/*_sunburst.csv')
    input_csv = csv_files[0] # first match
    df = pd.read_csv(input_csv)

    # Fill NaN values in the original DataFrame
    df_filled_na = fill_na_with_last_known(df.copy())

    # Sort the DataFrame by hierarchy and volume
    df_filled_sorted = sort_sunburst_hierarchy(df_filled_na)

    # Undo the fill with the original values
    df_final = undo_fill_with_original(df_filled_sorted, df)

    # Save the sorted DataFrame to a new CSV file
    sorted_path = input_csv.replace('sunburst.csv', 'sunburst_sorted.csv')
    df_final.to_csv(sorted_path, index=False)

    if verbose:
        print(f'\nSunburst csv sorted by region hierarchy : \n')
        print(f'{df_final}\n')

    # Attempt to calculate top regions, collapsing as necessary
    criteria_met = False
    while not criteria_met:
        top_regions_df = calculate_top_regions(df_final, top_regions, percent_vol, verbose)

        if top_regions_df is not None and not top_regions_df.empty: # If top regions are found
            criteria_met = True

            # If a top region contributes to less than 1% of the total volume, remove it
            total_volume = df_final['Volume_(mm^3)'].sum()
            top_regions_df = top_regions_df[top_regions_df['Volume_(mm^3)'] / total_volume > 0.01]
            
            if verbose:
                print(f'\nTop regions meeting the volume criterion:\n')
                print(f'{top_regions_df}\n')

            # Initialize lists to hold the top region names and their aggregate volumes
            top_region_names_and_percent_vols = []

            # For each top region, get the region name and the aggregate volume
            for i, row in top_regions_df.iterrows():

                # Get the highest depth region name that is not NaN
                row_wo_volume = row[:10]
                region_name = row_wo_volume[::-1].dropna().iloc[0]

                # Calculate the percentage of the total volume the top regions represent
                aggregate_volume = row['Volume_(mm^3)']
                percent_vol = round(aggregate_volume / total_volume * 100)

                # Append the region name and the percentage volume to the list
                top_region_names_and_percent_vols.append(f'{region_name} ({percent_vol}%)')

            print(f'\n{total_volume=} mm^3')
            print(f'\n{top_region_names_and_percent_vols=}')

            # Save the top regions DataFrame to a new CSV file
            top_regions_path = input_csv.replace('sunburst.csv', 'sunburst_top_regions.csv')
            top_regions_df.to_csv(top_regions_path, index=False)
        else:
            # Attempt to collapse the hierarchy further
            df_final = collapse_hierarchy(df_final, verbose)

            if verbose:
                print(f'\nTop regions do not meet the volume criterion. Collapsing the hierarchy:\n')
                print(f'{df_final}\n')

            if df_final.empty:
                break  # Exit if no further collapsing is possible

    return top_region_names_and_percent_vols, total_volume




def main():
    args = parse_args()

    # Find cluster_* dirs in the current dir
    cluster_dirs = glob('cluster_*')

    # Generate dynamic column names based on args.top_regions
    column_names = ['Cluster'] + ['Volume'] + ['CoG'] + ['~Region'] + ['ID_Path'] + [f'Top_Region_{i+1}' for i in range(args.top_regions)]

    # Load the *cluster_info.txt file from the parent dir to get the cluster Centroid of Gravity (CoG)
    parent_dir = Path.cwd().parent
    cluster_info_files = glob(str(parent_dir / '*cluster_info.txt'))
    cluster_info_file = cluster_info_files[0] # first match
 
    # Load the cluster info file
    cluster_info_df = pd.read_csv(cluster_info_file, sep='\t', header=None)  # Assuming tab-separated values; adjust if different

    # Reverse the row order of only the first column (excluding the header)
    first_column_name = cluster_info_df.columns[0]
    reversed_data = cluster_info_df[first_column_name].iloc[1:].iloc[::-1].reset_index(drop=True) # Reverse the data 
    new_first_column = pd.concat([cluster_info_df[first_column_name].iloc[:1], reversed_data]) # Concat header w/ reversed data
    cluster_info_df[first_column_name] = new_first_column.values # Assign the new column back to the DataFrame

    # Get the CoG for each cluster (values in last three columns of each row)
    CoGs = cluster_info_df.iloc[:, -3:].values

    # Convert to this format: 'x,y,z'
    CoGs = [','.join(map(str, CoG)) for CoG in CoGs]

    # Create a dict w/ the first column as keys and the CoG as values
    cluster_CoGs = dict(zip(cluster_info_df[first_column_name], CoGs))

    # Create an empty DataFrame with the column names
    top_regions_and_percent_vols_df = pd.DataFrame(columns=column_names)

    # Load csv with CCFv3 info 
    ccfv3_info_df = pd.read_csv(Path(__file__).parent / 'CCFv3_info.csv')

    # For each cluster directory
    for cluster_dir in cluster_dirs:
        cluster_num = cluster_dir.split('_')[-1]  # Extract cluster number from directory name

        # Get the CoG string for the current cluster from the dictionary
        cog_string = cluster_CoGs.get(cluster_num) if cluster_CoGs.get(cluster_num) else "Not found"

        # Get the top regions and their percentage volumes for the current cluster
        top_regions_and_percent_vols, cluster_volume = get_top_regions_and_percent_vols(cluster_dir, args.top_regions, args.percent_vol, args.verbose)

        # Get the top region
        top_region = top_regions_and_percent_vols[0].split(' ')[0]

        # Lookup the general_regio and structure_id_path in the DataFrame using the abbreviation for the top region
        general_region = ccfv3_info_df.loc[ccfv3_info_df['abbreviation'] == top_region, 'general_region'].values
        id_path = ccfv3_info_df.loc[ccfv3_info_df['abbreviation'] == top_region, 'structure_id_path'].values

        # Since there could be multiple matches, we take the first one or a default value if not found
        general_region = general_region[0] if len(general_region) > 0 else "General region not found"
        id_path = id_path[0] if len(id_path) > 0 else "ID path not found"

        # Ensure the list has the exact number of top regions (pad with None if necessary)
        padded_top_regions = (list(top_regions_and_percent_vols) + [None] * args.top_regions)[:args.top_regions]

        # Prepare the row data, including placeholders for 'Volume', 'CoG', '~Region', and top regions
        row_data = [cluster_num, cluster_volume, cog_string, general_region, id_path] + padded_top_regions   

        # Ensure column_names matches the structure of row_data
        column_names = ['Cluster', 'Volume', 'CoG', '~Region', 'ID_Path'] + [f'Top_Region_{i+1}' for i in range(args.top_regions)]

        # Create a temporary DataFrame for the current cluster's data
        temp_df = pd.DataFrame([row_data], columns=column_names)

        # Concatenate the temporary DataFrame with the main DataFrame
        top_regions_and_percent_vols_df = pd.concat([top_regions_and_percent_vols_df, temp_df], ignore_index=True)

    # Sort the DataFrame by the 'ID_Path' column in descending order
    top_regions_and_percent_vols_df = top_regions_and_percent_vols_df.sort_values(by='ID_Path', ascending=False)

    # Drop the 'ID_Path' column
    top_regions_and_percent_vols_df = top_regions_and_percent_vols_df.drop(columns=['ID_Path'])

    # Convert the 'Volume' column to 4 decimal places
    top_regions_and_percent_vols_df['Volume'] = top_regions_and_percent_vols_df['Volume'].round(4)

    # Define the path for the new Excel file
    excel_file_path = Path(__file__).parent / 'top_regions_and_percent_vols.xlsx'

    print(f'\n The top regions and their percentage volumes for each cluster: \n')
    print(f'\n{top_regions_and_percent_vols_df}\n')

    # Load csv with RGB values 
    sunburst_RGBs_df = pd.read_csv(Path(__file__).parent / 'sunburst_RGBs.csv', header=None)

    # Parse the dataframe to get a dictionary of region names and their corresponding RGB values
    rgb_values = {}
    for index, row in sunburst_RGBs_df.iterrows():
        # Format is 'region_name: rgb(r,g,b)'
        region, rgb = row[0].split(': rgb(')
        r, g, b = map(int, rgb.strip(')').split(','))
        rgb_values[region] = (r, g, b)

    # Create an Excel workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Write each row of the DataFrame to the worksheet and color cells for the top regions
    for region in dataframe_to_rows(top_regions_and_percent_vols_df, index=False, header=True):
        ws.append(region)

        for i in range(args.top_regions):

            # Find the region name without the percentage to match the RGB values
            top_region_column_num = 4 + i

            region_key = None
            if region[top_region_column_num] is not None:
                region_key = region[top_region_column_num].split(' ')[0]

            # Apply the color to the cell if it matches one of the RGB values
            if region_key in rgb_values:
                # Convert the RGB to a hex string
                rgb = rgb_values[region_key]
                hex_color = "{:02x}{:02x}{:02x}".format(rgb[0], rgb[1], rgb[2])
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                top_region_column_num = 5 + i
                ws.cell(row=ws.max_row, column=top_region_column_num).fill = fill
                ws.cell(row=ws.max_row, column=top_region_column_num).border = thin_border
            elif region_key is None:
                hex_color = "{:02x}{:02x}{:02x}".format(100, 100, 100) # Grey
                fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')
                top_region_column_num = 5 + i
                ws.cell(row=ws.max_row, column=top_region_column_num).fill = fill
                ws.cell(row=ws.max_row, column=top_region_column_num).border = thin_border

    # Apply a thin border style to cells with content
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:  # If the cell has content
                cell.border = thin_border
                cell.font = Font(name='Arial', size=11)

    # Apply the font to the header row
    header_font = Font(name='Arial', bold=True)
    for cell in ws['1:1']:
        cell.font = header_font

    # Insert a new row at the top
    ws.insert_rows(1)

    # Insert a new column at the left
    ws.insert_cols(1)            

    # Adjust the column width to fit the content
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if max_length > 0:
            adjusted_width = max_length + 2  # Add 2 for a little extra padding
            column_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Fill cells in first column with white
    for cell in ws['A']:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in ninth column with white
    column = 6 + args.top_regions
    for cell in ws[get_column_letter(column)]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in first row with white
    for cell in ws[1]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Fill cells in last row with white
    num_rows = top_regions_and_percent_vols_df.shape[0] + 3
    for cell in ws[num_rows]:
        cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # Center align the content
    for row in ws.iter_rows(min_row=1, min_col=1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the workbook to a file
    excel_file_path = 'valid_clusters_table.xlsx'
    wb.save(excel_file_path)
    print(f"Excel file saved at {excel_file_path}")


if __name__ == '__main__':
    main()